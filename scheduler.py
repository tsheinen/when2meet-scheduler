from z3 import *
from itertools import product
import json
from math import ceil
from sys import exit, argv
import argparse

from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import logging as log

log.basicConfig(format='%(levelname)s:%(message)s', level=log.INFO)


parser = argparse.ArgumentParser(description='Generate unique schedules from a when2meet instance')
parser.add_argument('url',metavar="URL",type=str)
parser.add_argument('-l', '--lower-bound', default=4, type=int)
parser.add_argument('-u', '--upper-bound', default=16, type=int)
parser.add_argument('-f', '--filter', type=str, nargs="*")
parser.add_argument('-n', '--num', default=1, type=int)

args = parser.parse_args()

def get_data(url):

    options = Options()
    options.headless = True
    driver = webdriver.Firefox(options=options)
    driver.get(url)


    # see PeopleIDs in when2meet js console
    pids = [x for x in driver.execute_script("return PeopleIDs;")]
    # see PeopleNames in when2meet js console
    names = [x for x in driver.execute_script("return PeopleNames;")]

    people = {k:v for k,v in zip(pids, names)}
    # see AvailableAtSlot in when2meet js console
    availability = driver.execute_script("return AvailableAtSlot;")
    driver.quit()
    return (people, availability)


log.info(f"...scraping data from when2meet: {args.url}")
people, availability = get_data(args.url)
people = {k:v for k,v in people.items() if v not in args.filter}
pids, names = zip(*people.items())
log.info("...scraped data, solving for constraints")


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

"""
header rows and columns
TODO this should be dynamic
"""
days = ["Monday", "Tuesday", "Wednesday", "Thursday","Friday","Saturday"]
times = ["10:00","10:15", "10:30", "10:45", "11:00", "11:15", "11:30", "11:45", "12:00", "12:15", "12:30", "12:45", "1:00", "1:15", "1:30", "1:45", "2:00", "2:15", "2:30", "2:45", "3:00", "3:15", "3:30", "3:45", "4:00", "4:15", "4:30", "4:45", "5:00", "5:15", "5:30", "5:45"]

if len(days) * len(times) != len(availability):
    print(f"You have a different number of headers/rows ({len(days) * len(times)}) than availability slots ({len(availability)})")
    exit(1)
timeslots = {}

s = Optimize()


"""
Initial constraint setup -- each timeslot may be assigned to any person which has marked that slot as available
"""
for (index_day, day), (index_time, time) in product(enumerate(days), enumerate(times)):
    index = index_day * 32 + index_time
    if len(availability[index]) > 0:
        key = day + "_" + time
        """
        Create a symbolic integer and add a constraint that is is equal to any ID which we have marked as available in that time slot
        """
        timeslots[key] = Int(key)
        s.add(Or([timeslots[key] == x for x in availability[index]]))

"""
Adds constraints that [length]-size slices are assigned to the same person
"""
def add_contiguous_constraint(length):
    for i in chunks(list(product(days, times)),length):
        keys = [x + "_" + y for x,y in i]
        constraints = []
        def add_pair(x,y):
            if keys[x] in timeslots and keys[y] in timeslots:
                constraints.append(timeslots[keys[x]] == timeslots[keys[y]])

        for x in range(0,length-1):
            add_pair(x,x+1)
        s.add(And(constraints))

"""
Add soft constraints that [length]-size slices are assigned to the same person
"""
def add_soft_contiguous_constraint(length, weight):
    for i in chunks(list(product(days, times)),length):
        keys = [x + "_" + y for x,y in i]
        constraints = []
        def add_pair(x,y):
            if keys[x] in timeslots and keys[y] in timeslots:
                constraints.append(timeslots[keys[x]] == timeslots[keys[y]])

        for x in range(0,length-1):
            add_pair(x,x+1)
        s.add_soft(And(constraints),weight=weight)

"""
Sets an upper bound on the number of timeslots which may be assigned
"""
def restrict_max_slots(num):
    for pid in pids:
        """
        PbLe is a Z3 function which allows you to assert that a given number of booleans are true
        In this case we're asserting that no more than [num] assignments are equal to any given pid
        """
        s.add_soft(PbLe([(x == pid,1) for x in timeslots.values()],num),100)

"""
Sets a lower bound on the number of timeslots which may be assigned
"""
def restrict_min_slots(num):
    for pid in pids:
        s.add_soft(PbGe([(x == pid,1) for x in timeslots.values()],num),100)

add_contiguous_constraint(4)
add_soft_contiguous_constraint(8,10)
add_soft_contiguous_constraint(16,15)

restrict_max_slots(args.upper_bound)
restrict_min_slots(args.lower_bound)

from openpyxl import Workbook, styles

count = 0
wb = Workbook()  
while s.check() == sat:
    log.info("...satisfied constraint, adding sheet")
    sheet = wb.create_sheet(f"Assignment {count}", count) 
    count += 1
    model = s.model()


    """
    this doesn't do much useful here but if we wanted to generate multiple valid schedules this changes the constraints such that the current model isn't valid
    """
    s.add(Or([v != model[v].as_long() for v in timeslots.values()]))

    # set of visually distinct colors to use as fill color for people cells
    colors = ['e6194b', '3cb44b', 'ffe119', '4363d8', 'f58231', '911eb4', '46f0f0', 'f032e6', 'bcf60c', 'fabebe', '008080', 'e6beff', '9a6324', 'fffac8', '800000', 'aaffc3', '808000', 'ffd8b1', '000075', '808080']

    # we've generated a valid schedule, lets put it in an excel sheet now
    for k,v in timeslots.items():
        row = 2 + times.index(k.split("_")[1])
        col = 2 + days.index(k.split("_")[0])
        val = people[model[v].as_long()]
        cell = sheet.cell(row=row, column=col, value=val)
        cell.fill = styles.PatternFill(fill_type="solid", start_color=colors[names.index(val)])


    header_font = styles.Font(bold=True)
    cornflower_blue_fill = styles.PatternFill(fill_type="solid", start_color='6495ED')
    dark_gray_fill = styles.PatternFill(fill_type="solid", start_color='a9a9a9')


    # populate times in first column
    for index, time in enumerate(times):
        cell = sheet.cell(row=index + 2, column=1, value=time)
        cell.font = header_font
        cell.fill = dark_gray_fill

    # populate days (aka header) in first row
    for index, day in enumerate(days):
        cell = sheet.cell(row=1, column=index + 2, value=day)
        cell.fill = cornflower_blue_fill
        cell.font = header_font
    if count == args.num:
        break
if count != args.num:
    log.error(f"Couldn't generate {args.num} unique assignments")
wb.save("availability.xlsx") 